import os
import json
import openpyxl
import pandas as pd
import numpy as np
from openpyxl import load_workbook


class SystemCall:
    def __init__(self, name, call_type, modifiable_fields, condition_num=0, condition_areas=None, modifiable_num=0):
        self.name = name
        self.call_type = call_type
        self.condition_num = condition_num
        self.condition_num_max = condition_num
        self.condition_num_min = condition_num
        self.modifiable_num = modifiable_num
        self.modifiable_fields = modifiable_fields
        self.condition_areas = condition_areas if condition_areas else []

    def merge_modifiable_fields(self, new_areas):
        self.modifiable_fields = list(set(self.modifiable_fields + new_areas))

    def __str__(self):
        return (f"System Call: {self.name}, Type: {self.call_type}, "
                f"Modifiable Fields: {sorted(self.modifiable_fields)}, "
                f"Condition Fields: {sorted(self.condition_areas)}")

class SystemCallManager:
    def __init__(self):
        self.calls = []

    def add_system_call(self, system_call):
        existing_call = self.find_system_call(system_call.name)
        if existing_call:
            existing_call.merge_modifiable_fields(system_call.modifiable_fields)
            # pick the maxnium 
            if existing_call.condition_num_min > system_call.condition_num:
               existing_call.condition_num_min = system_call.condition_num
            if existing_call.condition_num_max < system_call.condition_num:
               existing_call.condition_num_max = system_call.condition_num
            if(len(existing_call.modifiable_fields)!=0):
                existing_call.modifiable_fields.sort()
                existing_call.modifiable_num = len(existing_call.modifiable_fields)
        else:
            self.calls.append(system_call)

    def find_system_call(self, name):
        for call in self.calls:
            if call.name == name:
                return call
        return None
    
    def sort_calls(self):
        # 对calls进行排序，名字中包含"create"的排在后面
        self.calls.sort(key=lambda call: 'create' in call.name.lower())

    def export_to_xlsx(self, filename):
        
        data = [
            {
                'Name': call.name,
                'Type': call.call_type,
                'Modifiable Fields': ','.join(map(str, call.modifiable_fields)),
                'Modifiable Num': int(call.modifiable_num),
                'Condition Num(Max/Min)': str(int(call.condition_num_max))+'/'+str(int(call.condition_num_min)),
                'Condition Fields': ','.join(map(str, call.condition_areas))
            }
            for call in self.calls
        ]
        df = pd.DataFrame(data)
        df.to_excel(filename, index=False, engine='openpyxl')

    def import_from_xlsx(self, filename):
        df = pd.read_excel(filename, engine='openpyxl')
        for _, row in df.iterrows():
            # Modifiable Fields
            areas = list(map(int, row['Modifiable Fields'].split(','))) if pd.notna(row['Modifiable Fields']) else []

            # Condition Fields
            if pd.notna(row['Condition Fields']):
                if isinstance(row['Condition Fields'], str):
                    condition_areas = list(map(int, row['Condition Fields'].split(',')))
                elif isinstance(row['Condition Fields'], int):
                    condition_areas = [row['Condition Fields']]
                else:
                    condition_areas = []
            else:
                condition_areas = []

            # Create a new SystemCall object
            system_call = SystemCall(row['Name'], row['Type'], areas, condition_areas)
            self.add_system_call(system_call)

    def display_all_calls(self):
        for call in self.calls:
            if 'create' in call.name.lower():
                call.condition_num += 1
            print(call)


def convert_to_excel(recorded_data, output_file):
    """
    Convert the recorded JSON data to an Excel file.

    Parameters:
    recorded_data (dict): The recorded data.
    output_file (str): The path to the output Excel file.
    """
    # Initialize an empty DataFrame
    df = pd.DataFrame()

    for subdir, data_list in recorded_data.items():
        for data in data_list:
            # Convert each sublist to a DataFrame and set 'subdir' as a column
            temp_df = pd.DataFrame(data)
            temp_df['subdir'] = subdir
            temp_df['type'] = subdir.split("_")[0].upper()
            # Append to the main DataFrame
            df = df.append(temp_df, ignore_index=True)

    # Reorder columns to have 'subdir' as the first column
    columns = ['subdir']+ ['type'] + [col for col in df.columns if col != 'subdir' and col != 'type']
    df = df[columns]
    # Save to Excel file
    df.to_excel(output_file, index=True)
    return df

def analyze_type_modifiable_locations(data):
    df_sorted = df.sort_values(by=['type', 'offset', 'value_controllable'], ascending=[True, True, False])
    df_unique = df.drop_duplicates(subset=['offset', 'type'])
    #print(df_unique)
    # Count total occurrences of each type
    type_counts = df_unique['type'].value_counts()

    # Count occurrences where 'value_controllable' and 'has_constraints' are True
    controllable_counts = df_unique[df_unique['value_controllable']]['type'].value_counts()
    constraints_counts = df_unique[df_unique['has_constraints']]['type'].value_counts()

    # Merging the counts into a single DataFrame for plotting
    merged_counts = pd.DataFrame({
        'Total Count': type_counts,
        'Value Controllable Count': controllable_counts,
        'Has Constraints Count': constraints_counts
    }).fillna(0)

    # Plotting
    fig, ax = plt.subplots(figsize=(12, 8))
    merged_counts.plot(kind='bar', ax=ax)
    ax.set_title('Comparison of Type Counts')
    ax.set_xlabel('Type')
    ax.set_ylabel('Count')
    plt.savefig("modifiable_locations_counts.png")




def type_analysis_of_modifiable_locations(df):
    df_sorted = df.sort_values(by=['type', 'offset', 'value_controllable'], ascending=[True, True, False])
    df = calculate_offset_summary(df)
    return df

def calculate_offset_summary(df):
    """
    Calculate a summary of offset values for each type, including minimum offset, maximum offset,
    and the minimum and maximum gaps between any two offsets (not necessarily consecutive).

    Parameters:
    df (DataFrame): The DataFrame containing the data.

    Returns:
    DataFrame: A DataFrame with the summary for each type.
    """
    summary = {}
    for t in df['type'].unique():
        type_df = df[df['type'] == t]
        type_df = type_df.drop_duplicates(subset=['offset', 'type'])
        offsets = type_df['offset'].unique()
        value_controllable = type_df['value_controllable']
        has_constraints = type_df['has_constraints']
        controllable_count=0
        no_constraints_count=0
        fully_control_count = 0
        #print(value_controllable)
        for i in range(len(value_controllable)):
            if value_controllable.iloc[i] == True:
                controllable_count=controllable_count+1
            if has_constraints.iloc[i] == False:
                no_constraints_count=no_constraints_count+1
            if value_controllable.iloc[i]==True and has_constraints.iloc[i]==False:
                fully_control_count = fully_control_count +1
        offsets.sort()
        min_gap = float('inf')
        max_gap = float('-inf')
        for i in range(len(offsets) - 1):
            for j in range(i + 1, len(offsets)):
                gap = offsets[j] - offsets[i]
                min_gap = min(min_gap, gap)
                max_gap = max(max_gap, gap)
        #print(t)
        summary[t] = {
            'modifiable_count':len(offsets),
            'controllable_count':controllable_count,
            'no_constraint':no_constraints_count,
            "fully_controllable_count":fully_control_count,
            'min_offset': min(offsets),
            'max_offset': max(offsets),
            'min_gap': min_gap,
            'max_gap': max_gap
        }
    df = pd.DataFrame.from_dict(summary, orient='index')
    #print(df)
    return df.astype(int)




def traverse_and_record_json_data(directory):
    """
    Traverse all subdirectories in the given directory, read JSON files,
    and record data as specified.

    Parameters:
    directory (str): The root directory to start traversal.

    Returns:
    dict: A dictionary containing subdirectory names and their JSON data.
    """
    data = {}

    # Traverse the directory
    for root, dirs, files in os.walk(directory):
        for subdir in dirs:
            subdirectory_path = os.path.join(root, subdir)
            data[subdir] = []
            subdir_data=[]
            # Traverse each subdirectory for JSON files
            for file in os.listdir(subdirectory_path):
                if file.endswith('.json'):
                    json_path = os.path.join(subdirectory_path, file)
                    
                    # Read and record JSON data
                    with open(json_path, 'r') as f:
                        
                        json_data = json.load(f)
                        if json_data['name'] != "lazy_alloc1":
                            continue
                        if subdir_data == []:
                            for key, value in json_data.items():
                                if key.startswith("writable location"):
                                    flag = True
                                    if value.get("constraints")==[]:
                                        flag = False
                                    else:
                                        flag = len(value.get("constraints"))
                                    subdir_data.append({"name":json_data["name"],"size":json_data["size"],"width":int(value.get("width"))/8,"offset":int(value.get("offset_in_mo"))/2,"value_controllable":value.get("value_controllable"),"can_value_be_forged_id": value.get("can_value_be_forged_id"),"has_constraints":flag})
                        for key, value in json_data.items():
                                if key.startswith("writable location"):
                                    is_coverd = False
                                    for i in range(len(subdir_data)):
                                        if subdir_data[i]["offset"] == value.get("offset_in_mo"):
                                            is_coverd = True
                                            flag = True
                                            if value.get("constraints")==[]:
                                                flag = False
                                                if subdir_data[i]["has_constraints"] != False:
                                                    subdir_data[i]["has_constraints"] = False
                                                    #print(subdirectory_path)
                                                    #print("Change has_constraints!!!")
                                            if value.get("value_controllable") == True:
                                                if subdir_data[i]["value_controllable"]==False:
                                                    subdir_data[i]["value_controllable"]=True
                                                    #print("Change value controllable!!!")
                                    if is_coverd == False:
                                        flag = True
                                        #print("Uncoverd location!")
                                        if value.get("constraints")==[]:
                                            flag = False
                                        else:
                                            flag = len(value.get("constraints"))
                                        subdir_data.append({"name":json_data["name"],"size":json_data["size"],"width":int(value.get("width"))/8,"offset":int(value.get("offset_in_mo"))/2,"value_controllable":value.get("value_controllable"),"can_value_be_forged_id": value.get("can_value_be_forged_id"),"has_constraints":flag})

                        # for key, value in json_data.items():
                        #     if key == "value_controllable" and value == "true"
                        #     if key.startswith("modifiable location"):
                        #         if value.get("constraints")==[]:
                        #             count+=1
                        
            data[subdir].append(subdir_data)

    return data


os.chdir('/home/klee/klee_src/symbolic-execution-experiment/results')
directory = 'test-info-output'
recorded_data = traverse_and_record_json_data(directory)
output_file = 'modifiable_location_each_syscall_analysis.xlsx'
df = convert_to_excel(recorded_data, output_file)


manager = SystemCallManager()
manager_can_be_forged = SystemCallManager()
manager_all = SystemCallManager()
workbook = load_workbook(filename="modifiable_location_each_syscall_analysis.xlsx")
sheet = workbook.active
for row in sheet.iter_rows(min_row=2, values_only=True):  # The first line is the title
    syscall_name, object_type, mo_name, offset, value_controllable, can_be_forged_id, condition_num = row[1], row[2], row[3], row[6], row[7], row[8],row[9]
    
    # if mo_name == "lazy_alloc1":
    #     if condition_num == False:
    #         syscall = manager.add_system_call(SystemCall(syscall_name, object_type, [offset], 0))
    #     else:
    #         syscall = manager.add_system_call(SystemCall(syscall_name, object_type, [offset], condition_num))

    ## find those fields that is value controllable
    if mo_name == "lazy_alloc1" and value_controllable == True:
        if condition_num == False:
            if "create" in syscall_name:
                syscall = manager.add_system_call(SystemCall(syscall_name, object_type, [offset], 1))
            else:
                syscall = manager.add_system_call(SystemCall(syscall_name, object_type, [offset], 0))
        else:
            if "create" in syscall_name:
                syscall = manager.add_system_call(SystemCall(syscall_name, object_type, [offset], condition_num+1))
            else:
                syscall = manager.add_system_call(SystemCall(syscall_name, object_type, [offset], condition_num))
    if mo_name == "lazy_alloc1" and value_controllable == True and can_be_forged_id:
        if condition_num == False:
            if "create" in syscall_name:
                syscall = manager_can_be_forged.add_system_call(SystemCall(syscall_name, object_type, [offset], 1))
            else:
                syscall = manager_can_be_forged.add_system_call(SystemCall(syscall_name, object_type, [offset], 0))
        else:
            if "create" in syscall_name:
                syscall = manager_can_be_forged.add_system_call(SystemCall(syscall_name, object_type, [offset], condition_num+1))
            else:
                syscall = manager_can_be_forged.add_system_call(SystemCall(syscall_name, object_type, [offset], condition_num))
    if mo_name == "lazy_alloc1":
        if condition_num == False:
            if "create" in syscall_name:
                syscall = manager_all.add_system_call(SystemCall(syscall_name, object_type, [offset], 1))
            else:
                syscall = manager_all.add_system_call(SystemCall(syscall_name, object_type, [offset], 0))
        else:
            if "create" in syscall_name:
                syscall = manager_all.add_system_call(SystemCall(syscall_name, object_type, [offset], condition_num+1))
            else:
                syscall = manager_all.add_system_call(SystemCall(syscall_name, object_type, [offset], condition_num))
#manager.display_all_calls()
manager_all.export_to_xlsx('M1_vulnerable_system_calls.xlsx')
print("The modifibale fields (M1) result is saved in M1_vulnerable_system_calls.xlsx")
manager.export_to_xlsx('M2_vulnerable_system_calls.xlsx')
print("The modifibale fields (M2) result is saved in M2_vulnerable_system_calls.xlsx")
manager_can_be_forged.export_to_xlsx('M3_vulnerable_system_calls.xlsx')
print("The modifibale fields (M3) result is saved in M3_vulnerable_system_calls.xlsx")




